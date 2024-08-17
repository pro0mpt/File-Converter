import tkinter as tk
from tkinter import filedialog, messagebox
import os
from docx import Document
from pdf2docx import Converter
from openpyxl import load_workbook

# Function to handle file conversion (Add other conversions as needed)
def convert_file(file_path, convert_to):
    output_file = ""
    
    if file_path.endswith('.pdf') and convert_to == 'docx':
        # Convert PDF to DOCX
        output_file = file_path.replace('.pdf', '.docx')
        cv = Converter(file_path)
        cv.convert(output_file)
        cv.close()
        
    elif file_path.endswith('.docx') and convert_to == 'pdf':
        # DOCX to PDF (example logic, needs external library like ReportLab)
        output_file = file_path.replace('.docx', '.pdf')
        # Conversion logic here...
        
    elif file_path.endswith('.xlsx') and convert_to == 'csv':
        # XLSX to CSV conversion
        output_file = file_path.replace('.xlsx', '.csv')
        wb = load_workbook(file_path)
        ws = wb.active
        with open(output_file, 'w') as f:
            for row in ws.iter_rows(values_only=True):
                f.write(','.join(map(str, row)) + '\n')
                
    else:
        messagebox.showerror("Error", "Unsupported file format or conversion")
        return None
    
    return output_file

# Main application window
def open_converter():
    def upload_file():
        file_path = filedialog.askopenfilename()
        if file_path:
            entry_file.delete(0, tk.END)
            entry_file.insert(0, file_path)

    def start_conversion():
        file_path = entry_file.get()
        convert_to = dropdown_var.get()

        if not file_path or not convert_to:
            messagebox.showwarning("Warning", "Please select a file and conversion format")
            return

        output_file = convert_file(file_path, convert_to)
        if output_file:
            messagebox.showinfo("Success", f"File converted successfully! Saved as {output_file}")
    
    # GUI Setup
    window = tk.Tk()
    window.title("Document Converter")
    window.geometry("400x200")
    window.resizable(False, False)

    # File Upload Section
    label_file = tk.Label(window, text="Select a document:")
    label_file.pack(pady=10)
    
    entry_file = tk.Entry(window, width=50)
    entry_file.pack(pady=5)

    btn_upload = tk.Button(window, text="Browse", command=upload_file)
    btn_upload.pack(pady=5)

    # Dropdown for Conversion Formats
    dropdown_var = tk.StringVar(window)
    dropdown_var.set("Select Format")

    formats = ["docx", "pdf", "xlsx", "csv"]  # Add as needed
    dropdown = tk.OptionMenu(window, dropdown_var, *formats)
    dropdown.pack(pady=5)

    # Convert Button
    btn_convert = tk.Button(window, text="Convert", command=start_conversion)
    btn_convert.pack(pady=10)

    window.mainloop()

# Launch the converter
if __name__ == "__main__":
    open_converter()
